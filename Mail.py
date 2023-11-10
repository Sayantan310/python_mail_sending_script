import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.mime.text import MIMEText
from openpyxl.reader.excel import *



def send_email(sender_email, sender_password, receiver_email, subject, message, attachment_path):
    # Create a multipart message
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject

    # Add message body
    msg.attach(MIMEText(message, 'plain'))

    # Open the file in bynary
    with open(attachment_path, 'rb') as attachment:
        # Add file as application/octet-stream
        # Email client can usually download this automatically as attachment
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())

    # Encode file in ASCII characters to send by email
    encoders.encode_base64(part)

    # Add header as key/value pair to attachment part
    part.add_header(
        'Content-Disposition',
        f'attachment; filename= {attachment_path}'
    )

    # Add attachment to message and convert message to string
    msg.attach(part)
    text = msg.as_string()

    # Send email using SMTP
    with smtplib.SMTP('smtp.gmail.com', 587) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, receiver_email, text)

# Example usage
sender_email = 'sayantanbiswas310@gmail.com'
sender_password = 'cgknzocbpfsygrjz'
# receiver_email = 'sayantanb@live.in'
# subject = f'Candidate Seeking Interview Opportunity at {}'  # Subject of the mail
attachment_path = 'Sayantan_Biswas_CV.pdf'  # Replace with the actual file path #CV


my_file = load_workbook("Partners_2023_07_26_1.xlsx")  #Replace it with the Data Set name
s = my_file['Sheet1']  # dont forget to replace it
row = s.max_row
my_name = "Sayantan Biswas"
ph = "7003513962"
experience = "1"
linkdin = "https://www.linkedin.com/in/sayantanbiswas98"
designation = "Associate System Engineer"
my_company = "Accenture"
specialization = "Microsoft Azure DevOps "
unsent = []
sent = []

# FOR DATASET 1
for i in range(2,row + 1):
    r_name = None  #Recruiter Name
    company_Name = None
    receiver_email = None
    r_name = s.cell(i,1).value
    if s.cell(i,2).value is not None:
        r_name = r_name + " " + s.cell(i,2).value
    if s.cell(i,3).value is None or s.cell(i,4).value is None:
        if s.cell(i,3).value is None:
            print("Can't send mail to",r_name,"as email address is missing")
            unsent.append(r_name)
            receiver_email = None
            r_name = None
            company_Name = None
            continue
        if s.cell(i,4).value is None:
            print("Can't send mail to",r_name,"as Company Name is missing")
            unsent.append(r_name)
            receiver_email = None
            r_name = None
            company_Name = None
            continue
    else:
        sent.append(r_name)
        company_Name = s.cell(i,4).value
        receiver_email = s.cell(i,3).value
        subject = f'Seeking Opportunity at {company_Name}'
        message = f'''Dear {r_name},

I am reaching out to convey my keen interest in joining {company_Name} and to inquire about potential interview opportunities based on my skills. With {experience}+ years of experience as an {designation} at {my_company}, specializing in {specialization}, I am confident in my ability to make a valuable impact. PFA my detailed resume for your reference.

Having researched {company_Name}'s exceptional reputation, I am eager to contribute my skills and expertise to your team.
I would be grateful for the chance to discuss my qualifications further in an interview and provide additional insights into how my skills can contribute to {company_Name}'s ongoing success.

I look forward to the opportunity to discuss further for a role in your organization. Should you require any additional information or if there is a more appropriate contact person within your organization, please let me know.

I appreciate your time and consideration. I eagerly await your response.

Sincerely,
{my_name}
Phone No.: {ph}
Email ID: {sender_email}
LinkedIn ID: {linkdin}'''  # Email body context
        try:
            send_email(sender_email, sender_password, receiver_email, subject, message, attachment_path)
            print("Email send to :",r_name,",Associative Email Address :",receiver_email)
        except Exception as e:
            print(e)
            continue
print("Successful Email sent list : ",sent,"COUNT = ",len(sent))
print("Unsuccessful Email sent list",unsent,"COUNT = ",len(unsent))

# FOR DATASET 2
# for i in range(2, row + 1):
#     r_name = None  # Recruiter Name
#     company_Name = None
#     receiver_email = None
#     if s.cell(i, 1).value is not None and s.cell(i,4).value is not None and s.cell(i,9).value is not None:
#         r_name = s.cell(i,1).value
#         company_Name = s.cell(i,9).value
#         receiver_email = s.cell(i,4).value
#         subject = f'Seeking Opportunity at {company_Name}'
#         message = f'''Dear {r_name},
#
# I am reaching out to convey my keen interest in joining {company_Name} and to inquire about potential interview opportunities based on my skills. With {experience}+ years of experience as an {designation} at {my_company}, specializing in {specialization}, I am confident in my ability to make a valuable impact. PFA my detailed resume for your reference.
#
# Having researched {company_Name}'s exceptional reputation, I am eager to contribute my skills and expertise to your team.
# I would be grateful for the chance to discuss my qualifications further in an interview and provide additional insights into how my skills can contribute to {company_Name}'s ongoing success.
#
# I look forward to the opportunity to discuss further for a role in your organization. Should you require any additional information or if there is a more appropriate contact person within your organization, please let me know.
#
# I appreciate your time and consideration. I eagerly await your response.
#
# Sincerely,
# {my_name}
# Phone No.: {ph}
# Email ID: {sender_email}
# LinkedIn ID: {linkdin}'''  # Email body context
#         try:
#             send_email(sender_email, sender_password, receiver_email, subject, message, attachment_path)
#             print(f"Email send to : {r_name}, Associative Email Address : {receiver_email}, Index : {i}")
#             sent.append(r_name)
#         except Exception as e:
#             print(e)
#             unsent.append(r_name)
#             continue
#     else:
#         if s.cell(i,1).value is None:
#             print(f"Index : {i} - Recruiter name is missing can't send email")
#             unsent.append(s.cell(i,1).value)
#             continue
#         elif s.cell(i,4).value is None:
#             print(f"Index : {i} - Recruiter email address is missing can't send email")
#             unsent.append(s.cell(i,1).value)
#             continue
#         else:
#             print(f"Index : {i} - Recruiters Company name is missing can't send email")
#             unsent.append(s.cell(i, 1).value)
#             continue
#
# print("Successful Email sent list : ", sent, "COUNT = ", len(sent))
# print("Unsuccessful Email sent list", unsent, "COUNT = ", len(unsent))